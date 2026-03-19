from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

# Column color coding: A=shared, B=shared, C-E=Vendor A, F-H=Vendor B, I-K=comparison/comments
FILL_VENDOR_A = PatternFill(start_color="DAE3F3", end_color="DAE3F3", fill_type="solid")  # light blue
FILL_VENDOR_B = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # light orange
FILL_NEUTRAL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # light gray

from src.processing.comparison_builder import build_comparison_rows_for_scope


def _auto_fit_columns(ws) -> None:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # type: ignore[attr-defined]
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue
        ws.column_dimensions[column].width = min(max_length + 2, 60)


def _apply_comparison_column_colors(ws, max_row: int) -> None:
    """Apply Vendor A / Vendor B / neutral background to columns for easy differentiation."""
    # A-B: shared (neutral), C-E: Vendor A, F-H: Vendor B, I-K: comparison/comments (neutral)
    for row in range(1, max_row + 1):
        for col in range(1, 12):  # columns 1-11 (A-K)
            cell = ws.cell(row=row, column=col)
            if col <= 2:
                cell.fill = FILL_NEUTRAL
            elif col <= 5:
                cell.fill = FILL_VENDOR_A
            elif col <= 8:
                cell.fill = FILL_VENDOR_B
            else:
                cell.fill = FILL_NEUTRAL


def write_excel_workbook(
    output_path: Path,
    matches_with_deltas: pd.DataFrame,
    unmatched: pd.DataFrame,
    audit_df: pd.DataFrame,
    vendor_a_sheets: Dict[str, pd.DataFrame],
    vendor_b_sheets: Dict[str, pd.DataFrame],
    validation_ok: bool,
    unmatched_diagnostics: Optional[pd.DataFrame] = None,
    sheet_pairs: Optional[List[Tuple[str, str, str]]] = None,
) -> None:
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Summary sheet: comparison totals + counts
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Vendor comparison"
    summary["A1"].font = Font(bold=True, size=14)

    def _sum_total(sheets: Dict[str, pd.DataFrame]) -> float:
        total = 0.0
        for df in sheets.values():
            if "total_price" in df.columns:
                total += df["total_price"].fillna(0.0).sum()
        return round(total, 2)

    total_a = _sum_total(vendor_a_sheets)
    total_b = _sum_total(vendor_b_sheets)
    diff_abs = round(total_b - total_a, 2) if total_a is not None and total_b is not None else None
    diff_pct = round((diff_abs / total_a * 100), 1) if total_a and total_a != 0 and diff_abs is not None else None

    summary["A3"] = "Total (Vendor A)"
    summary["B3"] = total_a
    summary["A4"] = "Total (Vendor B)"
    summary["B4"] = total_b
    summary["A5"] = "Difference (B − A)"
    summary["B5"] = diff_abs if diff_abs is not None else ""
    summary["A6"] = "Difference (%)"
    summary["B6"] = f"{diff_pct}%" if diff_pct is not None else ""

    summary["A8"] = "Lines compared (matched)"
    summary["B8"] = int(len(matches_with_deltas))
    summary["A9"] = "Lines not matched (only one vendor quoted)"
    summary["B9"] = int(len(unmatched))

    summary["A11"] = "Note: Totals can differ — not all vendors quote every item. Use this file for line-by-line comparison."
    summary["A11"].font = Font(italic=True)

    # By-section comparison: use sheet_pairs so "Mechanics" and "ATS Mechanics v2" show as one section
    summary["A13"] = "By section"
    summary["A13"].font = Font(bold=True)
    summary["A14"] = "Section"
    summary["B14"] = "Total (A)"
    summary["C14"] = "Total (B)"
    summary["D14"] = "Difference"
    summary["A14"].font = Font(bold=True)
    summary["B14"].font = Font(bold=True)
    summary["C14"].font = Font(bold=True)
    summary["D14"].font = Font(bold=True)
    row = 15
    if sheet_pairs:
        for scope, a_sheet, b_sheet in sheet_pairs:
            a_df = vendor_a_sheets.get(a_sheet)
            b_df = vendor_b_sheets.get(b_sheet)
            sa = round(a_df["total_price"].fillna(0).sum(), 2) if a_df is not None and "total_price" in a_df.columns else 0.0
            sb = round(b_df["total_price"].fillna(0).sum(), 2) if b_df is not None and "total_price" in b_df.columns else 0.0
            summary.cell(row=row, column=1, value=scope)
            summary.cell(row=row, column=2, value=sa)
            summary.cell(row=row, column=3, value=sb)
            summary.cell(row=row, column=4, value=round(sb - sa, 2))
            row += 1
    else:
        scopes_for_summary = sorted(set(list(vendor_a_sheets.keys()) + list(vendor_b_sheets.keys())))
        for scope in scopes_for_summary:
            a_df = vendor_a_sheets.get(scope)
            b_df = vendor_b_sheets.get(scope)
            sa = round(a_df["total_price"].fillna(0).sum(), 2) if a_df is not None and "total_price" in a_df.columns else 0.0
            sb = round(b_df["total_price"].fillna(0).sum(), 2) if b_df is not None and "total_price" in b_df.columns else 0.0
            summary.cell(row=row, column=1, value=scope)
            summary.cell(row=row, column=2, value=sa)
            summary.cell(row=row, column=3, value=sb)
            summary.cell(row=row, column=4, value=round(sb - sa, 2))
            row += 1

    # Per-scope comparison sheets
    scopes = sorted(set(matches_with_deltas["scope_category"].dropna().tolist()))

    for scope in scopes:
        sheet_name = f"{scope} Comparison"
        ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name limit

        headers = [
            "Item ref",
            "Item name",
            "Qty (A)",
            "Unit price (A)",
            "Total (A)",
            "Qty (B)",
            "Unit price (B)",
            "Total (B)",
            "Price difference ($)",
            "Price difference (%)",
            "Comments",
        ]
        ws.append(headers)

        rows = build_comparison_rows_for_scope(
            scope,
            matches_with_deltas,
            vendor_a_sheets,
            vendor_b_sheets,
        )
        for row in rows:
            ws.append(row)

        max_row = ws.max_row
        _apply_comparison_column_colors(ws, max_row)

        # Simple conditional formatting on Price Delta (%) column (J)
        pct_col = "J"
        if max_row >= 2:
            cell_range = f"{pct_col}2:{pct_col}{max_row}"
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="greaterThan",
                    formula=["0.05"],
                    fill=PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    ),
                ),
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="between",
                    formula=["0.01", "0.05"],
                    fill=PatternFill(
                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                    ),
                ),
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="lessThanOrEqual",
                    formula=["0.01"],
                    fill=PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    ),
                ),
            )

        _auto_fit_columns(ws)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Template-based writer (Comparisonv1.xlsm)
# ---------------------------------------------------------------------------
# We keep the original writer above for reference, but the function name is
# intentionally redefined here so this implementation overrides it.
def write_excel_workbook(  # type: ignore[override]
    output_path: Path,
    matches_with_deltas: pd.DataFrame,
    unmatched: pd.DataFrame,
    audit_df: pd.DataFrame,
    vendor_a_sheets: Dict[str, pd.DataFrame],
    vendor_b_sheets: Dict[str, pd.DataFrame],
    validation_ok: bool,
    unmatched_diagnostics: Optional[pd.DataFrame] = None,
    sheet_pairs: Optional[List[Tuple[str, str, str]]] = None,
) -> None:
    """
    Generate an output workbook that matches `Comparisonv1.xlsm` layout/formatting
    and keeps its built-in conditional formatting/features.

    Implementation strategy:
    - Copy the template .xlsm file
    - Fill `Customer`, `Vendor1`, and `Vendor2` input tables
    - Leave existing formulas/conditional formatting in `Comparison` intact
    """
    import shutil
    from openpyxl import load_workbook
    from pandas import isna

    # v1 is an .xlsm template; ensure the output stays .xlsm.
    if output_path.suffix.lower() != ".xlsm":
        raise ValueError(
            f"Template-based output requires .xlsm, got: {output_path.suffix}"
        )

    base_dir = Path(__file__).resolve().parents[2]  # vendor-comparison-engine/
    template_path = base_dir / "Comparisonv1.xlsm"
    if not template_path.exists():
        raise FileNotFoundError(f"Excel template not found: {template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path, keep_vba=True, data_only=False)

    # Required template sheets
    ws_customer = wb["Customer"]
    ws_v1 = wb["Vendor1"]
    ws_v2 = wb["Vendor2"]

    # Data starts at row 5 (header at row 4)
    row_start = 5
    # Comparison sheet has formulas up to at least row 107; safe upper bound.
    max_rows = 107 - row_start + 1

    def _safe_float(v) -> float | None:
        if v is None or isna(v):
            return None
        try:
            f = float(v)
        except Exception:
            return None
        return f

    def _get_unit_price(row: pd.Series) -> float | None:
        for k in ("unit_price_sum", "unit_price_hardware", "unit_price_service"):
            if k in row.index:
                f = _safe_float(row.get(k))
                if f is not None:
                    return f
        return None

    def _get_total_price(row: pd.Series) -> float | None:
        if "total_price" in row.index:
            return _safe_float(row.get("total_price"))
        # Fallback to unit_price * qty if possible
        up = _get_unit_price(row)
        qty = _safe_float(row.get("qty"))
        if up is None or qty is None:
            return None
        return up * qty

    def _row_has_money(row: pd.Series) -> bool:
        tp = _get_total_price(row)
        up = _get_unit_price(row)
        if tp is not None and tp != 0:
            return True
        if up is not None and up != 0:
            return True
        return False

    def _get_ci(df: pd.DataFrame, substrs: List[str]) -> str | None:
        """Find a column whose name includes any of the substrings (case-insensitive)."""
        for col in df.columns:
            col_l = str(col).lower()
            if any(s.lower() in col_l for s in substrs):
                return col
        return None

    def _parse_item_id_sort(item_id: object) -> tuple:
        s = "" if item_id is None else str(item_id).strip()
        if not s:
            return (10, s)
        parts = s.split(".")
        if all(p.isdigit() for p in parts):
            return (0, tuple(int(p) for p in parts))
        # fallback: try numbers inside
        nums = []
        for p in re.findall(r"\d+", s):
            nums.append(int(p))
        return (1, tuple(nums), s)

    # Build line-items keyed by (scope, item_id).
    # Each entry contains vendor A/B rows where available.
    entries: Dict[tuple, Dict[str, object]] = {}

    def _add_matched(scope: str, mrow: pd.Series) -> None:
        a_sheet = mrow.get("vendor_a_sheet")
        b_sheet = mrow.get("vendor_b_sheet")
        a_idx = mrow.get("vendor_a_idx")
        b_idx = mrow.get("vendor_b_idx")
        if isna(a_idx) or isna(b_idx):
            return
        a_idx_i = int(a_idx)
        b_idx_i = int(b_idx)

        a_df = vendor_a_sheets.get(a_sheet)
        b_df = vendor_b_sheets.get(b_sheet)
        if a_df is None or b_df is None:
            return
        a_row = a_df.loc[a_idx_i]
        b_row = b_df.loc[b_idx_i]

        item_id = a_row.get("item_id") or b_row.get("item_id")
        if item_id is None or str(item_id).strip() == "":
            return
        if not _row_has_money(a_row) and not _row_has_money(b_row):
            return

        # Baseline qty is vendor A qty (qty check will flag mismatches vs vendor B).
        qty_a = _safe_float(a_row.get("qty"))
        qty_b = _safe_float(b_row.get("qty"))
        baseline_qty = qty_a if qty_a is not None else (qty_b if qty_b is not None else 0.0)

        opt_col_a = _get_ci(a_df, ["opt"])  # vendor-specific option column name
        opt_col_b = _get_ci(b_df, ["opt"])

        entries[(scope, str(item_id))] = {
            "scope": scope,
            "item_id": str(item_id),
            "baseline_qty": baseline_qty,
            "qty_a": qty_a if qty_a is not None else baseline_qty,
            "qty_b": qty_b if qty_b is not None else baseline_qty,
            "unit_price_a": _get_unit_price(a_row) or 0.0,
            "unit_price_b": _get_unit_price(b_row) or 0.0,
            "unit_a": a_row.get("unit") if "unit" in a_row.index else "",
            "unit_b": b_row.get("unit") if "unit" in b_row.index else "",
            "option_a": (a_row.get(opt_col_a) if opt_col_a else ""),
            "option_b": (b_row.get(opt_col_b) if opt_col_b else ""),
            "comment_a": a_row.get("comments") if "comments" in a_row.index else "",
            "comment_b": b_row.get("comments") if "comments" in b_row.index else "",
        }

    # NOTE: python doesn't know about `re` here; import inside function.
    import re

    if not matches_with_deltas.empty:
        for _, m in matches_with_deltas.iterrows():
            scope = m.get("scope_category")
            if scope is None or str(scope).strip() == "":
                continue
            _add_matched(str(scope), m)

    # Unmatched items: include those where one vendor has a row.
    if not unmatched.empty:
        for _, u in unmatched.iterrows():
            scope = u.get("scope_category")
            if scope is None or str(scope).strip() == "":
                continue

            a_sheet = u.get("vendor_a_sheet")
            b_sheet = u.get("vendor_b_sheet")
            a_idx = u.get("vendor_a_idx")
            b_idx = u.get("vendor_b_idx")

            a_df = vendor_a_sheets.get(a_sheet)
            b_df = vendor_b_sheets.get(b_sheet)

            a_row = None
            b_row = None
            if not isna(a_idx) and a_df is not None:
                a_row = a_df.loc[int(a_idx)]
            if not isna(b_idx) and b_df is not None:
                b_row = b_df.loc[int(b_idx)]

            if a_row is None and b_row is None:
                continue

            item_id = (a_row.get("item_id") if a_row is not None else None) or (
                b_row.get("item_id") if b_row is not None else None
            )
            if item_id is None or str(item_id).strip() == "":
                continue

            # Skip pure-blank/phantom rows (neither side has a price)
            if (a_row is not None and _row_has_money(a_row)) or (
                b_row is not None and _row_has_money(b_row)
            ):
                pass
            else:
                continue

            scope_str = str(scope)
            item_id_str = str(item_id)
            if (scope_str, item_id_str) in entries:
                continue

            qty_a = _safe_float(a_row.get("qty")) if a_row is not None else None
            qty_b = _safe_float(b_row.get("qty")) if b_row is not None else None
            baseline_qty = qty_a if qty_a is not None else (qty_b if qty_b is not None else 0.0)

            opt_col_a = _get_ci(a_df, ["opt"]) if a_df is not None else None
            opt_col_b = _get_ci(b_df, ["opt"]) if b_df is not None else None

            entries[(scope_str, item_id_str)] = {
                "scope": scope_str,
                "item_id": item_id_str,
                "baseline_qty": baseline_qty,
                "qty_a": qty_a if qty_a is not None else baseline_qty,
                "qty_b": qty_b if qty_b is not None else baseline_qty,
                "unit_price_a": _get_unit_price(a_row) or 0.0 if a_row is not None else 0.0,
                "unit_price_b": _get_unit_price(b_row) or 0.0 if b_row is not None else 0.0,
                "unit_a": a_row.get("unit") if a_row is not None and "unit" in a_row.index else "",
                "unit_b": b_row.get("unit") if b_row is not None and "unit" in b_row.index else "",
                "option_a": (a_row.get(opt_col_a) if (a_row is not None and opt_col_a) else "") if a_row is not None else "",
                "option_b": (b_row.get(opt_col_b) if (b_row is not None and opt_col_b) else "") if b_row is not None else "",
                "comment_a": a_row.get("comments") if a_row is not None and "comments" in a_row.index else "",
                "comment_b": b_row.get("comments") if b_row is not None and "comments" in b_row.index else "",
            }

    # Sort entries: scope then parsed item id.
    entry_list = list(entries.values())
    entry_list.sort(key=lambda e: (str(e.get("scope", "")), _parse_item_id_sort(e.get("item_id"))))

    # Fill template tables.
    # Column mapping based on template headers (row 4):
    # Customer: A=Line Item, B=Scope, C=Unit Price, D=Option, E=Qty., F=Unit, G=Total Price, H=Comment
    # Vendor1 : C=Unit Price, D=Option, E=Amount, F=Unit, H=Comment (A/B are driven by Customer formulas)
    # Vendor2 : same columns
    for i, e in enumerate(entry_list[:max_rows]):
        r = row_start + i

        scope = e.get("scope", "")
        item_id = e.get("item_id", "")
        baseline_qty = _safe_float(e.get("baseline_qty")) or 0.0

        unit_price_a = _safe_float(e.get("unit_price_a")) or 0.0
        unit_price_b = _safe_float(e.get("unit_price_b")) or 0.0
        qty_a = _safe_float(e.get("qty_a")) or baseline_qty
        qty_b = _safe_float(e.get("qty_b")) or baseline_qty

        # Customer baseline unit price: prefer vendor A, else vendor B
        customer_unit_price = unit_price_a if unit_price_a != 0 else unit_price_b

        customer_option = e.get("option_a") if e.get("option_a") not in (None, "") else e.get("option_b", "")
        customer_unit = e.get("unit_a") if e.get("unit_a") not in (None, "") else e.get("unit_b", "")
        customer_comment = e.get("comment_a") if e.get("comment_a") not in (None, "") else e.get("comment_b", "")

        ws_customer.cell(row=r, column=1, value=item_id)
        ws_customer.cell(row=r, column=2, value=scope)
        ws_customer.cell(row=r, column=3, value=customer_unit_price)
        ws_customer.cell(row=r, column=4, value=customer_option)
        ws_customer.cell(row=r, column=5, value=baseline_qty)
        ws_customer.cell(row=r, column=6, value=customer_unit)
        ws_customer.cell(row=r, column=7, value=(customer_unit_price * baseline_qty) if customer_unit_price else 0.0)
        ws_customer.cell(row=r, column=8, value=customer_comment)

        # Vendor A inputs
        ws_v1.cell(row=r, column=3, value=unit_price_a)
        ws_v1.cell(row=r, column=4, value=e.get("option_a", ""))
        ws_v1.cell(row=r, column=5, value=qty_a)
        ws_v1.cell(row=r, column=6, value=e.get("unit_a", ""))
        ws_v1.cell(row=r, column=8, value=e.get("comment_a", ""))

        # Vendor B inputs
        ws_v2.cell(row=r, column=3, value=unit_price_b)
        ws_v2.cell(row=r, column=4, value=e.get("option_b", ""))
        ws_v2.cell(row=r, column=5, value=qty_b)
        ws_v2.cell(row=r, column=6, value=e.get("unit_b", ""))
        ws_v2.cell(row=r, column=8, value=e.get("comment_b", ""))

    wb.save(output_path)

