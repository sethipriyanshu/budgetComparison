from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

import openpyxl
import pandas as pd


NON_DATA_SHEET_KEYWORDS = {"summary", "cover", "toc", "index"}
MIN_DATA_ROWS = 5


@dataclass
class SheetData:
    """Container for raw sheet data and minimal metadata."""

    sheet_name: str
    dataframe: pd.DataFrame
    vendor_id: str
    source_path: Path


def _is_data_sheet(sheet_name: str, df: pd.DataFrame) -> bool:
    """Heuristically determine if a sheet contains tabular quote data."""
    name = sheet_name.strip().lower()
    if any(keyword in name for keyword in NON_DATA_SHEET_KEYWORDS):
        return False

    # Rough heuristic: count non-empty rows
    non_empty_rows = (df.notna().any(axis=1)).sum()
    return non_empty_rows >= MIN_DATA_ROWS


def _workbook_to_dataframes(
    path: Path, vendor_id: str
) -> Dict[str, SheetData]:
    """Load an .xlsx workbook and convert sheets to DataFrames."""
    wb = openpyxl.load_workbook(path, data_only=True)

    sheet_map: Dict[str, SheetData] = {}
    for ws in wb.worksheets:
        # Unmerge all merged cells to avoid header detection issues.
        if ws.merged_cells.ranges:
            merged_ranges: List[openpyxl.worksheet.cell_range.CellRange] = list(
                ws.merged_cells.ranges
            )
            for cell_range in merged_ranges:
                ws.unmerge_cells(range_string=str(cell_range))

        data: List[List[object]] = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        if not data:
            continue

        df = pd.DataFrame(data)
        if not _is_data_sheet(ws.title, df):
            continue

        sheet_map[ws.title] = SheetData(
            sheet_name=ws.title,
            dataframe=df,
            vendor_id=vendor_id,
            source_path=path,
        )

    return sheet_map


def read_vendor_workbook(path: str | Path, vendor_id: str) -> Dict[str, SheetData]:
    """
    Read a vendor .xlsx file and return a mapping of sheet name to SheetData.

    This function is intentionally strict about file existence and type to make
    CLI error handling straightforward.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Vendor file not found: {p}")
    if p.suffix.lower() not in {".xlsx"}:
        raise ValueError(f"Unsupported file type for {p.name}; expected .xlsx")

    return _workbook_to_dataframes(p, vendor_id=vendor_id)

